from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "VSR Comparison"

headers = [
    "Model",
    "Year / Venue",
    "Foundation Model",
    "Foundation Params (M)",
    "Trainable Params (M)",
    "Total Inference Params (M)",
    "Training Strategy",
    "Training GPU Type",
    "Training GPU Count",
    "Training Iterations",
    "Training Resolution",
    "Training Batch Size",
    "Training Data",
    "Degradation",
    "Optimizer (lr)",
    "Total GPU-hours / Time",
    "Includes REDS in train?",
    "A6000x2 Inference",
    "Source for training info",
    "Notes",
]

data = [
    # Non-diffusion baseline
    [
        "BasicVSR++",
        "CVPR 2022",
        "(none, recurrent CNN)",
        "—",
        "7.32",
        "7.32",
        "From scratch",
        "",
        "",
        "300K (supp)",
        "",
        "",
        "",
        "bicubic + others",
        "",
        "",
        "",
        "Yes (very light)",
        "arxiv 2104.13371",
        "Non-diffusion baseline; trained on REDS",
    ],
    # SD2.1-based (light, foundation ~868M)
    [
        "StableVSR (baseline)",
        "ECCV 2024",
        "claudiom4sir/StableVSR (SD2.1 variant)",
        "868",   # UNet 472 + VAE 55 + text 340
        "~207",
        "~1,075",
        "Adapter (ControlNet + TCM, UNet frozen)",
        "",
        "",
        "",
        "",
        "",
        "REDS",
        "bicubic",
        "",
        "~17 GB inference (paper)",
        "Yes (REDS-trained)",
        "Yes",
        "GitHub README / arxiv 2311.15908",
        "Our direct baseline; same foundation as Ours",
    ],
    [
        "Ours",
        "2026 (thesis)",
        "claudiom4sir/StableVSR (SD2.1 variant)",
        "868",
        "213.92",
        "1,082",
        "Adapter (ControlNet + dual-SFT, UNet frozen)",
        "NVIDIA RTX A6000",
        "2",
        "30,000",
        "512x512 (crop)",
        "",
        "REDS",
        "bicubic",
        "AdamW",
        "(fill in)",
        "Yes",
        "Yes (~13 GB inference)",
        "(internal)",
        "Dual-SFT with DT-CWT frequency separation",
    ],
    [
        "DGAF-VSR",
        "CVPR 2026 (arxiv 2511.16928)",
        "SD2.1 variant (block_out_channels=[256,512,512,1024])",
        "868",
        "277 (ControlNet)",
        "1,145",
        "Adapter (ControlNet, UNet frozen, OGWM+FTCM)",
        "Tesla V100-32G (inference confirmed)",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "Training details in Supplementary 8 (not available)",
        "(unknown)",
        "Yes (16-19 GB)",
        "README + arxiv (training in supp)",
        "Uses REDS4 + Vid4 + VideoLQ for testing",
    ],
    [
        "DLoRAL",
        "NeurIPS 2025 (arxiv 2506.15591)",
        "Stable Diffusion V2.1 + RAM-Swin",
        "1,800",
        "10.49 (LoRA r=4)",
        "1,810",
        "LoRA (C-LoRA + D-LoRA, frozen UNet)",
        "NVIDIA A100",
        "4",
        "(not specified)",
        "512x512",
        "16",
        "REDS + Pexels videos + LSDIR (image)",
        "RealESRGAN pipeline",
        "Adam (5e-5)",
        "(not specified)",
        "Yes (REDS for consistency stage)",
        "Yes",
        "Paper Sec 4.1",
        "REDS used in training (consistency stage)",
    ],
    [
        "UltraVSR",
        "ACM MM 2025 (arxiv 2505.19958)",
        "Stable Diffusion 2.1 + RAM-Swin",
        "1,800",
        "10.49 (LoRA r=4)",
        "1,810",
        "LoRA (VAE+UNet, RTS modules)",
        "",
        "",
        "",
        "512x512 (seq len 6)",
        "",
        "",
        "",
        "Adam (β2=0.99)",
        "",
        "",
        "OOM (Ours env)",
        "Paper",
        "LoRA on VAE + UNet; one-step diffusion",
    ],
    [
        "Upscale-A-Video",
        "CVPR 2024 (arxiv 2312.06640)",
        "SD x4 Upscaler (own multi-component)",
        "1,206",
        "(temporal layers + propagator)",
        "1,206",
        "Partial FT (temporal layers in UNet+VAE)",
        "",
        "",
        "",
        "",
        "",
        "WebVid10M (335K) + YouHQ-Train (38,576 videos)",
        "RealESRGAN-style",
        "",
        "",
        "(unknown)",
        "OOM (Ours env)",
        "Paper",
        "Flow-guided latent propagation (training-free)",
    ],
    # Wan2.1-based
    [
        "FlashVSR",
        "CVPR 2026 (arxiv 2510.12747)",
        "Wan 2.1-1.3B",
        "1,547",   # Wan diffusion 1420 + Wan VAE 127
        "1,880",  # LoRA r=384 + LQ proj + TCDec (large rank LoRA)
        "1,880",
        "LoRA r=384 (3-stage distillation)",
        "NVIDIA A100-80G",
        "32",
        "(stages 1-3 + TC Dec)",
        "768x1280 (89-frame clips)",
        "32",
        "VSR-120K (120k videos + 180k images)",
        "RealBasicVSR pipeline",
        "AdamW (1e-5, wd 0.01)",
        "~7 days total (2+1+2+2)",
        "Yes (REDS test reported in paper)",
        "Yes (11.1 GB peak)",
        "Paper Sec 4.1",
        "Block-sparse attention + streaming inference",
    ],
    # T2V-based (heavy)
    [
        "Vivid-VR",
        "ICLR 2026 (arxiv 2508.14483)",
        "CogVideoX1.5-5B + CogVLM2-llama3",
        "12,000",
        "~1,372 (controlnet + connectors)",
        "13,400",
        "Adapter (controlnet + dual-branch connector)",
        "NVIDIA H20-96G",
        "32",
        "30,000",
        "1024x1024",
        "1 per GPU (total 32)",
        "500K real + 100K generated videos + text",
        "(real-world degradations)",
        "AdamW (1e-4, cosine annealing)",
        "~6,000 GPU-hours",
        "No (1024² training, no REDS)",
        "Yes (~24 GB peak measured)",
        "Paper Sec 4.1",
        "Inference: 50 steps DPM solver, 1024² training",
    ],
    [
        "STAR",
        "ICCV 2025 (arxiv 2501.02976)",
        "I2VGen-XL (default) / CogVideoX-5B (alt)",
        "5,000+",
        "~2,041",
        "~7,000",
        "Adapter (ControlNet + LIEM)",
        "NVIDIA A100-80G",
        "8",
        "15,000",
        "720x1280 (32 frames)",
        "8",
        "OpenVid-1M subset (~200K text-video pairs)",
        "Real-ESRGAN + video compression",
        "AdamW (5e-5)",
        "(not specified)",
        "Tests REDS30 (not REDS4)",
        "Yes (~39 GB for small input)",
        "Paper Sec 4.1",
        "Init from VEnhancer weights; LIEM enhancement",
    ],
    [
        "DOVE",
        "NeurIPS 2025 (arxiv 2505.16239)",
        "CogVideoX1.5 (5B, full FT)",
        "5,000+ (replaced)",
        "~7,500 (full fine-tune)",
        "7,500",
        "Full FT (2-stage)",
        "NVIDIA A800-80G",
        "4",
        "10,500 (10K + 500)",
        "320x640 (25 frames)",
        "8 total",
        "HQ-VSR (2,055 videos) + DIV2K (900 images)",
        "RealBasicVSR (video) + Real-ESRGAN (image)",
        "AdamW (2e-5 stage1, 5e-6 stage2)",
        "(not specified)",
        "No (HQ-VSR / DIV2K)",
        "Yes (some configs)",
        "Paper Sec 4.1",
        "Empty text prompt; trained on 320x640",
    ],
    # From scratch
    [
        "SeedVR2",
        "ICLR 2026 (arxiv 2506.05301)",
        "own DiT (no separate foundation)",
        "—",
        "3,391",
        "3,391",
        "From scratch (adversarial post-training)",
        "NVIDIA H100-80G",
        "72",
        "(not specified)",
        "100 frames at 720p per batch",
        "(seq+data parallel)",
        "(not specified)",
        "(real-world)",
        "(not specified)",
        "(not specified)",
        "(unknown)",
        "OOM (Ours env)",
        "Paper / OpenReview",
        "One-step DiT; 72 H100 for training",
    ],
    [
        "MGLD-VSR",
        "ECCV 2024 (arxiv 2312.00853)",
        "Stable Diffusion 2.1",
        "1,290",
        "(motion-guided latent module)",
        "(unknown)",
        "Latent diffusion + motion guidance",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "A100 used for inference benchmark",
        "(unknown)",
        "Yes (likely)",
        "Paper",
        "Motion-guided latent sampling",
    ],
]

ws.append(headers)
for row in data:
    ws.append(row)

# Styling
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ours_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
baseline_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
oom_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
heavy_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
sd_light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Identify rows by foundation type for coloring
foundation_colors = {
    "claudiom4sir/StableVSR (SD2.1 variant)": sd_light_fill,
    "SD2.1 variant (block_out_channels=[256,512,512,1024])": sd_light_fill,
    "Stable Diffusion V2.1 + RAM-Swin": None,
    "Stable Diffusion 2.1 + RAM-Swin": None,
    "Stable Diffusion 2.1": None,
    "SD x4 Upscaler (own multi-component)": None,
    "Wan 2.1-1.3B": None,
    "CogVideoX1.5-5B + CogVLM2-llama3": heavy_fill,
    "I2VGen-XL (default) / CogVideoX-5B (alt)": heavy_fill,
    "CogVideoX1.5 (5B, full FT)": heavy_fill,
    "own DiT (no separate foundation)": heavy_fill,
    "(none, recurrent CNN)": baseline_fill,
}

for row_idx in range(2, len(data) + 2):
    model_name = ws.cell(row=row_idx, column=1).value
    foundation = ws.cell(row=row_idx, column=3).value
    oom_col = ws.cell(row=row_idx, column=18).value or ""

    fill = None
    if model_name == "Ours":
        fill = ours_fill
    elif "OOM" in oom_col:
        fill = oom_fill
    else:
        fill = foundation_colors.get(foundation)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        if fill is not None:
            cell.fill = fill

# Column widths
widths = {
    1: 18, 2: 22, 3: 30, 4: 14, 5: 18, 6: 18,
    7: 30, 8: 22, 9: 12, 10: 18, 11: 22, 12: 16,
    13: 32, 14: 24, 15: 24, 16: 22, 17: 24, 18: 20,
    19: 24, 20: 36,
}
for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 60
ws.freeze_panes = "B2"

# Add a legend sheet
legend = wb.create_sheet("Legend")
legend.append(["Color", "Meaning"])
legend.append(["", "Ours"])
legend.append(["", "SD-light category (~868M foundation, comparable to Ours)"])
legend.append(["", "Heavy T2V/from-scratch foundation (≥5B)"])
legend.append(["", "OOM in our A6000x2 environment"])
legend.append(["", "Non-diffusion baseline"])

legend.cell(row=2, column=1).fill = ours_fill
legend.cell(row=3, column=1).fill = sd_light_fill
legend.cell(row=4, column=1).fill = heavy_fill
legend.cell(row=5, column=1).fill = oom_fill
legend.cell(row=6, column=1).fill = baseline_fill

legend.cell(row=1, column=1).font = header_font
legend.cell(row=1, column=2).font = header_font
legend.cell(row=1, column=1).fill = header_fill
legend.cell(row=1, column=2).fill = header_fill
for row_idx in range(1, 7):
    for col_idx in range(1, 3):
        legend.cell(row=row_idx, column=col_idx).border = border
        legend.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

legend.column_dimensions['A'].width = 8
legend.column_dimensions['B'].width = 60

# Add a Notes sheet
notes = wb.create_sheet("Notes")
notes_lines = [
    ["#", "Note"],
    [1, "Foundation Params: For SD2.1-based models using claudiom4sir/StableVSR base, foundation = UNet (472M) + VAE (55M) + CLIP text encoder (340M) = 868M. This is smaller than stabilityai/stable-diffusion-2-1 (UNet 866M alone)."],
    [2, "DLoRAL uses REDS in TRAINING (consistency stage) — REDS4 evaluation may be biased."],
    [3, "STAR uses REDS30 (not REDS4) for testing, training degradation is RealESRGAN + video compression, not bicubic."],
    [4, "DOVE does NOT use REDS for training or testing. Trained on 320x640 resolution with RealBasicVSR/Real-ESRGAN degradation."],
    [5, "Vivid-VR trained on 1024x1024 resolution. REDS test (1280x720 non-square smaller) is out-of-distribution."],
    [6, "FlashVSR: 32x A100-80G for 7 days. Block-sparse attention from MIT-Han-Lab."],
    [7, "Several methods (DLoRAL/STAR/DOVE/FlashVSR/Vivid-VR) train with Real-ESRGAN or RealBasicVSR degradation; our REDS4 uses bicubic — fundamental degradation mismatch may explain low PSNR on bicubic test."],
    [8, "Ours, StableVSR (baseline) both use claudiom4sir/StableVSR foundation (smaller UNet: 472M vs SD2.1 standard 866M)."],
    [9, "DGAF-VSR training details are in Supplementary 8 (not available online); GitHub README confirms V100-32G for inference."],
    [10, "Vivid-VR's 6,000 GPU-hours on H20 ≈ ~11 months equivalent on a single A6000."],
]
for line in notes_lines:
    notes.append(line)
notes.cell(row=1, column=1).font = header_font
notes.cell(row=1, column=2).font = header_font
notes.cell(row=1, column=1).fill = header_fill
notes.cell(row=1, column=2).fill = header_fill
notes.column_dimensions['A'].width = 6
notes.column_dimensions['B'].width = 100
for row_idx in range(1, len(notes_lines) + 1):
    for col_idx in range(1, 3):
        notes.cell(row=row_idx, column=col_idx).border = border
        notes.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)

out = "/home/user/StableVSR/VSR_comparison_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
