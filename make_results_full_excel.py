"""Generate Excel comparison table for thesis defense.

Sheets:
  - REDS4 (from existing slide data)
  - Vid4
  - UDM10
  - SPMCS
  - Summary (per-metric winners per dataset)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "Method", "PSNR↑", "SSIM↑", "LPIPS↓", "DISTS↓",
    "MUSIQ↑", "CLIP-IQA↑", "NIQE↓", "tLPIPS↓", "tOF↓",
]

# All values measured/transcribed during defense prep.
# Ordering follows slide convention (Non-DM first, then DM by venue).
REDS4 = [
    ("BasicVSR++ (Non-DM)", 24.88, 0.730, 0.364, 0.179, 41.23, 0.265, 5.90, 38.26, 13.739),
    ("StableVSR (baseline)", 24.04, 0.690, 0.309, 0.164, 42.79, 0.237, 4.38, 41.11, 13.962),
    ("STAR",                 22.50, 0.618, 0.390, 0.177, 46.28, 0.214, 4.51, 21.25, 14.585),
    ("DOVE",                 23.46, 0.665, 0.351, 0.176, 43.69, 0.218, 4.78, 25.67, 15.563),
    ("DLoRAL",               22.33, 0.591, 0.265, 0.127, 66.06, 0.470, 3.09, 31.75, 15.404),
    ("Vivid-VR",             20.01, 0.536, 0.348, 0.136, 60.85, 0.299, 3.50, 23.33, 29.449),
    ("FlashVSR",             20.99, 0.569, 0.291, 0.129, 61.02, 0.288, 3.01, 17.17, 18.437),
    ("DGAF-VSR",             24.07, 0.694, 0.307, 0.161, 43.41, 0.242, 4.37, 40.12, 13.732),
    ("Ours",                 24.48, 0.691, 0.193, 0.088, 65.70, 0.386, 2.77, 15.25, 11.118),
    ("Ours+10k (30k total)", 24.24, 0.686, 0.189, 0.084, 66.10, 0.398, 2.92, 13.55, 11.098),
]

VID4 = [
    ("BasicVSR++ (Non-DM)", 26.26, 0.828, 0.189, 0.122, 61.50, 0.341, 5.04, 15.12, 0.489),
    ("StableVSR (baseline)", 22.98, 0.674, 0.185, 0.107, 67.22, 0.454, 3.19, 25.36, 0.751),
    ("DGAF-VSR",             23.29, 0.690, 0.177, 0.102, 67.96, 0.470, 3.10, 17.39, 0.688),
    ("Ours",                 22.64, 0.664, 0.194, 0.114, 66.15, 0.408, 3.32, 28.53, 0.920),
    ("DLoRAL",               20.85, 0.515, 0.296, 0.156, 68.49, 0.540, 3.19, 18.99, 2.100),
    ("FlashVSR",             19.65, 0.519, 0.263, 0.145, 70.09, 0.435, 3.12,  8.15, 1.843),
    ("STAR",                 18.18, 0.472, 0.323, 0.159, 68.30, 0.495, 4.35, 13.71, 1.634),
    ("Vivid-VR",             17.89, 0.397, 0.350, 0.184, 71.97, 0.524, 3.80, 15.24, 3.487),
    ("DOVE",                 22.72, 0.683, 0.213, 0.134, 65.13, 0.403, 4.49, 15.73, 1.603),
]

UDM10 = [
    ("BasicVSR++ (Non-DM)", 37.48, 0.956, 0.060, 0.053, 59.36, 0.443, 5.60,  5.55, 1.191),
    ("StableVSR (baseline)", 26.71, 0.834, 0.100, 0.055, 55.69, 0.362, 4.66,  4.42, 1.828),
    ("DGAF-VSR",             26.71, 0.835, 0.099, 0.054, 57.15, 0.380, 4.61,  2.99, 1.785),
    ("Ours",                 25.54, 0.811, 0.124, 0.066, 63.20, 0.447, 4.12, 14.48, 2.518),
    ("DLoRAL",               26.74, 0.769, 0.219, 0.135, 69.33, 0.654, 3.81, 18.19, 4.924),
    ("FlashVSR",             25.09, 0.765, 0.191, 0.096, 67.83, 0.470, 3.81, 10.60, 5.228),
    ("STAR",                 25.19, 0.772, 0.222, 0.119, 67.90, 0.516, 4.40, 10.33, 3.916),
    ("Vivid-VR",             23.93, 0.703, 0.275, 0.130, 68.89, 0.539, 3.95, 15.03, 6.390),
    ("DOVE",                 30.76, 0.879, 0.104, 0.065, 63.29, 0.456, 4.53,  6.78, 4.409),
]

SPMCS = [
    ("BasicVSR++ (Non-DM)", 21.94, 0.617, 0.187, 0.108, 62.48, 0.434, 5.17,  3.60, 0.345),
    ("StableVSR (baseline)", 19.42, 0.478, 0.196, 0.112, 69.98, 0.582, 3.28, 51.11, 0.652),
    ("DGAF-VSR",             20.06, 0.511, 0.178, 0.103, 67.70, 0.533, 3.61, 20.10, 0.461),
    ("Ours",                 19.94, 0.506, 0.183, 0.106, 67.22, 0.503, 3.70, 31.59, 0.575),
    ("DLoRAL",               21.08, 0.525, 0.255, 0.165, 70.55, 0.672, 3.65, 14.66, 1.498),
    ("FlashVSR",             20.23, 0.507, 0.216, 0.126, 71.26, 0.481, 3.39,  7.83, 1.397),
    ("STAR",                 19.27, 0.480, 0.294, 0.155, 69.55, 0.520, 4.12, 14.88, 1.138),
    ("Vivid-VR",             20.22, 0.494, 0.307, 0.154, 73.05, 0.560, 4.02,  7.67, 2.560),
    ("DOVE",                 22.58, 0.652, 0.168, 0.105, 70.58, 0.555, 4.31,  5.86, 1.527),
]

# Per metric direction (True = higher is better)
HIGHER_IS_BETTER = [True, True, False, False, True, True, False, False, False]


def best_indices(data, exclude_non_dm=False):
    """Return per-metric best row indices.

    Non-DM rows can dominate distortion metrics, so we also return DM-only
    bests to mark separately. `data` rows are tuples (name, *metrics).
    """
    overall = [None] * len(HIGHER_IS_BETTER)
    dm_only = [None] * len(HIGHER_IS_BETTER)
    for col, hib in enumerate(HIGHER_IS_BETTER):
        best_v, best_i = None, None
        best_dm_v, best_dm_i = None, None
        for i, row in enumerate(data):
            v = row[col + 1]
            is_non_dm = "Non-DM" in row[0]
            if best_v is None or (v > best_v if hib else v < best_v):
                best_v, best_i = v, i
            if not is_non_dm and (best_dm_v is None or (v > best_dm_v if hib else v < best_dm_v)):
                best_dm_v, best_dm_i = v, i
        overall[col] = best_i
        dm_only[col] = best_dm_i
    return overall, dm_only


def fill_sheet(ws, title, data):
    ws.title = title

    # Header
    ws.append(HEADERS)

    # Data rows
    for row in data:
        ws.append(list(row))

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    ours_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    best_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    dm_best_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    overall_best, dm_best = best_indices(data)

    for row_idx in range(2, len(data) + 2):
        method = ws.cell(row=row_idx, column=1).value
        is_ours = method.startswith("Ours")
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
            cell.border = border
            # numeric formatting
            if col_idx > 1:
                cell.number_format = "0.000" if col_idx in (3, 4, 5, 7, 10) else "0.00"

        if is_ours:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ours_fill

        data_row = row_idx - 2
        for metric_col, best_i in enumerate(overall_best):
            if data_row == best_i:
                c = ws.cell(row=row_idx, column=metric_col + 2)
                c.font = Font(bold=True)
                if not is_ours:
                    c.fill = best_fill
        for metric_col, dm_i in enumerate(dm_best):
            if data_row == dm_i and data_row != overall_best[metric_col]:
                c = ws.cell(row=row_idx, column=metric_col + 2)
                c.font = Font(bold=True)
                if not is_ours:
                    c.fill = dm_best_fill

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"


def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.append(["Dataset"] + HEADERS[1:])
    ws.append(["", "Method achieving the best on each metric (DM-only)"])
    ws.row_dimensions[2].height = 18

    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 3
    for ds_name, data in [("REDS4", REDS4), ("Vid4", VID4), ("UDM10", UDM10), ("SPMCS", SPMCS)]:
        _, dm_best = best_indices(data)
        row = [ds_name]
        for i in dm_best:
            row.append(data[i][0])
        ws.append(row)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=1).font = bold
        row_idx += 1

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "B2"


def build_legend(wb):
    ws = wb.create_sheet("Legend", 0)
    ws.title = "Legend"
    ws.append(["VSR Comparison — Thesis Defense"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Cell highlighting"])
    ws["A3"].font = Font(bold=True)
    ws.append(["", "Yellow", "Our method rows"])
    ws.cell(row=4, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.append(["", "Pink", "Best overall (bold)"])
    ws.cell(row=5, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.append(["", "Green", "Best among diffusion-based methods (DM-only, bold)"])
    ws.cell(row=6, column=2).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws.append([])
    ws.append(["Metrics"])
    ws["A8"].font = Font(bold=True)
    rows = [
        ("PSNR↑",  "Peak Signal-to-Noise Ratio (distortion / reconstruction)"),
        ("SSIM↑",  "Structural Similarity Index (distortion / reconstruction)"),
        ("LPIPS↓", "Learned Perceptual Image Patch Similarity (full-reference perceptual)"),
        ("DISTS↓", "Deep Image Structure and Texture Similarity (full-reference perceptual)"),
        ("MUSIQ↑", "Multi-Scale Image Quality (no-reference perceptual)"),
        ("CLIP-IQA↑", "CLIP-based Image Quality Assessment (no-reference perceptual)"),
        ("NIQE↓", "Natural Image Quality Evaluator (no-reference perceptual)"),
        ("tLPIPS↓", "Temporal LPIPS (temporal consistency, x1e3)"),
        ("tOF↓",   "Temporal Optical-Flow error (temporal consistency, x1e1)"),
    ]
    for k, v in rows:
        ws.append(["", k, v])
    ws.append([])
    ws.append(["Notes"])
    ws["A19"].font = Font(bold=True)
    notes = [
        "All models are evaluated with the same eval.py (torchmetrics, pyiqa, DISTS_pytorch).",
        "REDS4 is in-distribution for all REDS-trained methods. Vid4/UDM10/SPMCS are out-of-distribution.",
        "Non-DM (BasicVSR++) dominates pixel-fidelity metrics by design; DM methods optimize perceptual quality.",
        "DOVE uses different training data (HQ-VSR) and CogVideoX1.5-5B foundation; numbers are for reference.",
        "Vivid-VR (CogVideoX1.5-5B + CogVLM2) and FlashVSR (Wan2.1-1.3B) are heavier foundations.",
    ]
    for n in notes:
        ws.append(["", "•", n])

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80


def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for title, data in [("REDS4", REDS4), ("Vid4", VID4), ("UDM10", UDM10), ("SPMCS", SPMCS)]:
        ws = wb.create_sheet(title)
        fill_sheet(ws, title, data)

    build_summary(wb)
    build_legend(wb)

    out = "/home/user/StableVSR/results_comparison_full.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
