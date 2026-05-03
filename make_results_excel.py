"""Generate an Excel file summarizing dual-SFT evaluation results across datasets.

Usage:
    python make_results_excel.py
    # writes results_dualsft.xlsx in cwd

Requires: openpyxl  (pip install openpyxl)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (method, dataset) -> metrics
results = {
    ("Ours (dual-SFT)", "REDS4"):  {"PSNR": 24.48, "SSIM": 0.691, "LPIPS": 0.193, "DISTS": 0.088,
                                     "MUSIQ": 65.70, "CLIP-IQA": 0.386, "NIQE": 2.77,
                                     "tLPIPS": 15.25, "tOF": 11.12},
    ("Ours (dual-SFT)", "Vid4"):   {"PSNR": 22.64, "SSIM": 0.664, "LPIPS": 0.194, "DISTS": 0.114,
                                     "MUSIQ": 66.15, "CLIP-IQA": 0.408, "NIQE": 3.32,
                                     "tLPIPS": 28.53, "tOF": 0.920},
    ("Ours (dual-SFT)", "UDM10"):  {"PSNR": 25.54, "SSIM": 0.811, "LPIPS": 0.124, "DISTS": 0.066,
                                     "MUSIQ": 63.20, "CLIP-IQA": 0.447, "NIQE": 4.12,
                                     "tLPIPS": 14.48, "tOF": 2.518},
    ("Ours (dual-SFT)", "SPMCS"):  {"PSNR": 19.94, "SSIM": 0.506, "LPIPS": 0.183, "DISTS": 0.106,
                                     "MUSIQ": 67.22, "CLIP-IQA": 0.503, "NIQE": 3.70,
                                     "tLPIPS": 31.59, "tOF": 0.576},
    ("BasicVSR++",      "REDS4"):  {"PSNR": 24.88, "SSIM": 0.730, "LPIPS": 0.364, "DISTS": 0.179,
                                     "MUSIQ": 41.23, "CLIP-IQA": 0.265, "NIQE": 5.90,
                                     "tLPIPS": 38.26, "tOF": 13.739},
    ("BasicVSR++",      "Vid4"):   {"PSNR": 26.26, "SSIM": 0.828, "LPIPS": 0.189, "DISTS": 0.122,
                                     "MUSIQ": 61.50, "CLIP-IQA": 0.341, "NIQE": 5.04,
                                     "tLPIPS": 15.12, "tOF": 0.489},
    ("BasicVSR++",      "UDM10"):  {"PSNR": 37.48, "SSIM": 0.956, "LPIPS": 0.060, "DISTS": 0.053,
                                     "MUSIQ": 59.36, "CLIP-IQA": 0.443, "NIQE": 5.60,
                                     "tLPIPS": 5.55,  "tOF": 1.191},
    ("BasicVSR++",      "SPMCS"):  {"PSNR": 21.94, "SSIM": 0.617, "LPIPS": 0.187, "DISTS": 0.108,
                                     "MUSIQ": 62.48, "CLIP-IQA": 0.434, "NIQE": 5.17,
                                     "tLPIPS": 3.60,  "tOF": 0.345},
}

methods = ["BasicVSR++", "Ours (dual-SFT)"]
datasets = ["REDS4", "Vid4", "UDM10", "SPMCS"]

direction = {
    "PSNR": "↑", "SSIM": "↑",
    "LPIPS": "↓", "DISTS": "↓",
    "MUSIQ": "↑", "CLIP-IQA": "↑",
    "NIQE": "↓",
    "tLPIPS": "↓", "tOF": "↓",
}

groups = [
    ("Reconstruction", ["PSNR", "SSIM"]),
    ("Reference Perceptual", ["LPIPS", "DISTS"]),
    ("No-reference Perceptual", ["MUSIQ", "CLIP-IQA", "NIQE"]),
    ("Temporal", ["tLPIPS", "tOF"]),
]
all_metrics = [m for _, ms in groups for m in ms]

# ---------------------------------------------------------------- styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
group_fill = PatternFill("solid", fgColor="8EA9DB")
ds_fill = PatternFill("solid", fgColor="D9E1F2")
metric_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
best_font = Font(bold=True, color="C00000")

wb = Workbook()


# ============== Sheet 1: per-dataset comparison (rows = methods) ==============
ws = wb.active
ws.title = "By Dataset"

row = 1
for ds in datasets:
    # dataset header
    cell = ws.cell(row=row, column=1, value=ds)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=1 + len(all_metrics))
    row += 1

    # group header row
    ws.cell(row=row, column=1, value="Method").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=1).border = border
    col = 2
    for gname, gmetrics in groups:
        span = len(gmetrics)
        c = ws.cell(row=row, column=col, value=gname)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + span - 1)
        for k in range(col, col + span):
            ws.cell(row=row, column=k).border = border
        col += span
    row += 1

    # metric name row
    ws.cell(row=row, column=1, value="").fill = group_fill
    ws.cell(row=row, column=1).border = border
    for i, m in enumerate(all_metrics, start=2):
        c = ws.cell(row=row, column=i, value=f"{m} {direction[m]}")
        c.font = metric_font
        c.fill = group_fill
        c.alignment = center
        c.border = border
    row += 1

    # data rows + best detection
    method_rows = {}
    for method in methods:
        if (method, ds) not in results:
            continue
        method_rows[method] = row
        c = ws.cell(row=row, column=1, value=method)
        c.font = metric_font
        c.alignment = center
        c.border = border
        for i, m in enumerate(all_metrics, start=2):
            v = results[(method, ds)][m]
            cell = ws.cell(row=row, column=i, value=v)
            cell.alignment = center
            cell.border = border
            if m in ("PSNR", "MUSIQ", "tLPIPS"):
                cell.number_format = "0.00"
            elif m in ("NIQE", "tOF"):
                cell.number_format = "0.000"
            else:
                cell.number_format = "0.000"
        row += 1

    # bold best per metric (only if both methods present)
    if len(method_rows) > 1:
        for i, m in enumerate(all_metrics, start=2):
            vals = [(method, results[(method, ds)][m]) for method in method_rows]
            if direction[m] == "↑":
                best_method = max(vals, key=lambda x: x[1])[0]
            else:
                best_method = min(vals, key=lambda x: x[1])[0]
            ws.cell(row=method_rows[best_method], column=i).font = best_font

    row += 1  # blank line

# Column widths
ws.column_dimensions["A"].width = 18
for c in range(2, len(all_metrics) + 2):
    ws.column_dimensions[get_column_letter(c)].width = 11


# ============== Sheet 2: per-method (rows = datasets) ==============
ws2 = wb.create_sheet("By Method")
row = 1
for method in methods:
    cell = ws2.cell(row=row, column=1, value=method)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws2.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=1 + len(all_metrics))
    row += 1

    # group + metric headers
    ws2.cell(row=row, column=1, value="Dataset").font = header_font
    ws2.cell(row=row, column=1).fill = header_fill
    ws2.cell(row=row, column=1).alignment = center
    ws2.cell(row=row, column=1).border = border
    col = 2
    for gname, gmetrics in groups:
        span = len(gmetrics)
        c = ws2.cell(row=row, column=col, value=gname)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        if span > 1:
            ws2.merge_cells(start_row=row, start_column=col,
                            end_row=row, end_column=col + span - 1)
        for k in range(col, col + span):
            ws2.cell(row=row, column=k).border = border
        col += span
    row += 1

    ws2.cell(row=row, column=1, value="").fill = group_fill
    ws2.cell(row=row, column=1).border = border
    for i, m in enumerate(all_metrics, start=2):
        c = ws2.cell(row=row, column=i, value=f"{m} {direction[m]}")
        c.font = metric_font
        c.fill = group_fill
        c.alignment = center
        c.border = border
    row += 1

    for ds in datasets:
        if (method, ds) not in results:
            continue
        c = ws2.cell(row=row, column=1, value=ds)
        c.font = metric_font
        c.alignment = center
        c.border = border
        c.fill = ds_fill
        for i, m in enumerate(all_metrics, start=2):
            v = results[(method, ds)][m]
            cell = ws2.cell(row=row, column=i, value=v)
            cell.alignment = center
            cell.border = border
            if m in ("PSNR", "MUSIQ", "tLPIPS"):
                cell.number_format = "0.00"
            elif m in ("NIQE", "tOF"):
                cell.number_format = "0.000"
            else:
                cell.number_format = "0.000"
        row += 1

    row += 1

ws2.column_dimensions["A"].width = 14
for c in range(2, len(all_metrics) + 2):
    ws2.column_dimensions[get_column_letter(c)].width = 11


# ============== Sheet 3: Notes ==============
ws3 = wb.create_sheet("Notes")
notes = [
    "VSR Evaluation: Ours (dual-SFT) vs BasicVSR++",
    "",
    "Setup:",
    "  - Training: REDS only, ours 20K steps, BasicVSR++ checkpoint from authors (REDS).",
    "  - Inference: Ours uses 50 DDPM steps, bidirectional sampling.",
    "  - Eval script: /home/yjcho/StableVSR/eval.py (RGB PSNR; consistent across methods).",
    "",
    "Direction: ↑ higher is better, ↓ lower is better. Best in red bold (Sheet 1).",
    "",
    "Metrics:",
    "  Reconstruction:    PSNR, SSIM",
    "  Reference Percep.: LPIPS, DISTS  (deep-feature similarity to GT)",
    "  No-ref Percep.:    MUSIQ, CLIP-IQA, NIQE  (naturalness, no GT)",
    "  Temporal:          tLPIPS (perceptual frame-pair diff), tOF (optical-flow diff)",
    "",
    "Datasets (all BIx4 degradation):",
    "  REDS4   - 4 clips (000, 011, 015, 020) from REDS train split",
    "  Vid4    - calendar, city, foliage, walk",
    "  UDM10   - 10 clips",
    "  SPMCS   - 30 clips (Tao et al. 2017)",
    "",
    "Notes for paper:",
    "  - Ours wins on perceptual metrics (LPIPS, DISTS, MUSIQ, CLIP-IQA, NIQE).",
    "  - BasicVSR++ wins on PSNR/SSIM and temporal stability (tLPIPS, tOF).",
    "  - This perceptual-vs-fidelity trade-off matches StableVSR paper's findings.",
    "  - BasicVSR++ values for REDS4 / UDM10 not measured here; add if needed.",
]
for i, line in enumerate(notes, start=1):
    cell = ws3.cell(row=i, column=1, value=line)
    if i == 1:
        cell.font = Font(bold=True, size=12)
ws3.column_dimensions["A"].width = 90

wb.save("results_comparison.xlsx")
print("Wrote results_comparison.xlsx")

