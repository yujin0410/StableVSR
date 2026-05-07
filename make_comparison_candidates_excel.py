"""Generate comparison candidates Excel file.

Lists all VSR comparison candidate methods organized by category,
priority, and setup compatibility.

Usage:
    pip install openpyxl
    python make_comparison_candidates_excel.py
    # → comparison_candidates.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# (Method, Category, Direct compare, Cite, Setup, Priority, Notes)
ROWS = [
    # ── Already evaluated ──
    ("StableVSR (baseline)", "Diffusion", "DONE", "✓",
     "BIx4 ✓", "-", "Already evaluated. StableVSR baseline."),
    ("BasicVSR++",            "Feed-forward CNN", "DONE", "✓",
     "BIx4 ✓", "-", "Already evaluated. Strong feed-forward baseline."),
    ("Ours (dual-SFT)",       "Diffusion + Wavelet", "OURS", "-",
     "BIx4 ✓", "-", "Our method."),

    # ── Strong direct comparison ──
    ("DGAF-VSR (Rethinking VSR)", "Diffusion VSR", "★★★", "★★★",
     "BIx4 (TBC)", "🥇", "Dense Guidance Aligned Features. github: tszssong/DGAF-VSR"),
    ("RVRT", "Feed-forward Trans", "★★★", "★★★",
     "BIx4 ✓", "🥇", "Strong SOTA. github: JingyunLiang/RVRT"),
    ("PatchVSR", "Diffusion VSR", "★★★", "★★★",
     "BIx4 (TBC)", "🥇", "Patch-wise diffusion VSR"),

    # ── Recommended (if time) ──
    ("MGLD-VSR", "Diffusion VSR", "★★", "★★★",
     "BIx4 ✓", "🥈", "Local guidance diffusion. github: IanYeung/MGLD-VSR"),
    ("EDVR",     "Feed-forward CNN", "★★", "★★★",
     "BIx4 ✓", "🥈", "Classic VSR baseline (BasicSR repo)"),
    ("UltraVSR", "One-Step Diffusion", "★★", "★★",
     "TBC", "🥈", "Ultra-realistic, one-step efficiency"),
    ("DC-VSR",   "Diffusion VSR", "★★", "★★",
     "Standard", "🥈", "Spatially & Temporally Consistent"),
    ("SeeClear", "Diffusion VSR", "★★", "★★",
     "Standard", "🥈", "Semantic distillation"),
    ("VideoGigaGAN", "GAN VSR", "★★", "★★",
     "Standard", "🥈", "Strong GAN baseline"),

    # ── Cite as related work (different setup or focus) ──
    ("Upscale-A-Video", "Diffusion VSR (real-world)", "✗", "★★",
     "Real-world", "🥉", "Real-world degradation, different setup"),
    ("RealisVSR", "Diffusion VSR (real-world 4K)", "✗", "★★",
     "Real-world 4K", "🥉", "Different scale and setup"),
    ("STAR",     "Diffusion VSR (T2V-based)", "✗", "★★",
     "Real-world", "🥉", "T2V model for VSR, real-world"),
    ("Motion-Guided Latent Diffusion", "Diffusion VSR (real-world)", "✗", "★★",
     "Real-world", "🥉", "Motion-guided latent"),
    ("DOVE",     "One-step Real-world VSR", "✗", "★★",
     "Real-world", "🥉", "Efficient real-world"),
    ("OS-DiffVSR", "One-step Latent Diffusion", "✗", "★★",
     "Real-world", "🥉", "One-step real-world"),
    ("Improved Adversarial Diff Comp", "Diffusion VSR (real-world)", "✗", "★★",
     "Real-world", "🥉", "Real-world adversarial diffusion"),
    ("D2-VR",    "Degradation-Robust", "✗", "★★",
     "Robust degradation", "🥉", "Synergistic optimization"),
    ("Self-Sup ControlNet (Mamba)", "Real-world VSR", "✗", "★★",
     "Real-world", "🥉", "Spatio-Temporal Mamba"),

    # ── One-step / efficiency (different axis) ──
    ("InstaVSR",  "One-step Diffusion", "✗", "★★",
     "Standard", "🥉", "Efficiency focus"),
    ("DUO-VSR",   "One-step Diffusion", "✗", "★★",
     "Standard", "🥉", "Dual-Stream Distillation"),
    ("FlashVSR",  "Real-Time Streaming", "✗", "★★",
     "Streaming", "🥉", "Real-time diffusion"),
    ("Stream-DiffVSR", "Streaming Auto-Regressive", "✗", "★★",
     "Streaming", "🥉", "Low-latency streaming"),
    ("Asymmetric VAE One-Step", "One-step VSR", "✗", "★★",
     "Standard", "🥉", "VAE acceleration"),
    ("Towards Redundancy Reduction", "Efficient Diffusion", "✗", "★★",
     "Efficient", "🥉", "Efficiency optimization"),
    ("One-Step Detail-Rich VSR", "One-step Diffusion", "✗", "★★",
     "Standard", "🥉", "Detail-rich one-step"),
    ("TurboVSR",  "Efficient VSR", "✗", "★★",
     "TBC", "🥉", "Fantastic Video Upscalers"),
    ("SeedVR2",   "One-Step Restoration", "✗", "★★",
     "Generic", "🥉", "Diffusion adversarial post-training"),
    ("SimpleGVR", "Latent-Cascaded Baseline", "✗", "★",
     "Cascade", "🥉", "Simple baseline"),

    # ── Specialty / different problem ──
    ("FCVSR", "Frequency-aware (compressed)", "✗", "★★★",
     "Compressed VSR", "🥉", "Same frequency philosophy, different setup"),
    ("Fourier-Enhanced TecoGAN", "Frequency GAN", "△", "★★",
     "Standard", "🥉", "Frequency GAN-based"),
    ("Vivid-VR",  "Photorealistic Restoration", "✗", "★",
     "Generic", "🥉", "T2V distillation"),
    ("SeedVR",    "Generic Restoration", "✗", "★",
     "Generic", "🥉", "Generic VR with DiT"),
    ("DiTVR",     "Zero-Shot DiT VR", "✗", "★",
     "Zero-shot", "🥉", "Zero-shot diffusion transformer"),
    ("DiffIR2VR-Zero", "Zero-Shot Image→Video", "✗", "★",
     "Zero-shot", "🥉", "From image diffusion"),
    ("SparkVSR",  "Interactive Keyframe", "✗", "★",
     "Interactive", "🥉", "Sparse keyframe propagation"),
    ("UniMMVSR",  "Multi-Modal Cascade", "✗", "★",
     "Multi-modal", "🥉", "Cascaded multi-modal"),
    ("InfVSR",    "Length-Unrestricted", "✗", "★",
     "Long video", "🥉", "Breaking length limits"),
    ("DAM-VSR",   "Appearance-Motion Disentangle", "✗", "★",
     "Disentangle", "🥉", "Appearance/motion split"),
    ("LiftVSR",   "Image→Video VSR", "✗", "★",
     "TBC", "🥉", "Lift image diffusion to VSR"),
    ("STCDiT",    "Spatio-Temporally Consistent DiT", "★", "★★",
     "TBC", "🥉", "DiT-based VSR"),
    ("Aligning Global Semantics", "Latent Diffusion VSR", "★", "★★",
     "TBC", "🥉", "Semantic + texture alignment"),
    ("Improving Temporal Consistency", "Inference-time Zero-shot", "✗", "★",
     "Zero-shot", "🥉", "Image-based diffusion at inference"),

    # ── GAN-based ──
    ("DualX-VSR", "Real-world GAN (no MC)", "✗", "★",
     "Real-world", "🥉", "Dual axial transformer"),
    ("Collab Feedback Disc Prop", "Discriminative Prop", "✗", "★",
     "Standard", "🥉", "Feedback propagation"),
    ("Progressive Fusion GAN", "Realistic VSR GAN", "✗", "★",
     "Standard", "🥉", "Progressive fusion"),
    ("RBPGAN",    "Recurrent BP GAN", "✗", "★",
     "Standard", "🥉", "Recurrent back-projection"),
    ("GAN + Perceptual Losses", "Generic GAN VSR", "✗", "★",
     "Standard", "🥉", "Generic perceptual GAN"),

    # ── Skip (not appropriate) ──
    ("UltraGen",  "Video Generation", "✗✗", "✗",
     "Generation", "❌", "Not VSR (text-to-video etc.)"),
    ("EvTexture++", "Event-based VSR", "✗✗", "✗",
     "Event camera", "❌", "Different input modality"),
    ("FedVSR",    "Federated Learning", "✗✗", "✗",
     "Federated", "❌", "Different problem (distributed)"),
    ("WDASR",     "Medical (Cardiac MRI)", "✗✗", "✗",
     "Medical MRI", "❌", "Different domain"),
    ("Adversarial Diff Image SR", "Image SR", "✗✗", "✗",
     "Image, not video", "❌", "Image SR, not VSR"),
    ("Difiisr",   "Infrared Image SR", "✗✗", "✗",
     "Infrared image", "❌", "Different modality and image"),
    ("MambaVSR",  "Mamba SSM VSR", "△", "★",
     "BIx4 (TBC)", "❌", "Niche architecture, lower priority"),
    ("RepNet-VSR", "Reparameterization", "✗", "★",
     "Standard", "❌", "Already covered by BasicVSR++"),
]


HEADERS = ["Method", "Category", "Direct compare", "Cite",
           "Setup compat.", "Priority", "Notes"]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Candidates"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    # Priority color fills
    fills = {
        "🥇": PatternFill("solid", fgColor="C6EFCE"),  # green
        "🥈": PatternFill("solid", fgColor="FFEB9C"),  # yellow
        "🥉": PatternFill("solid", fgColor="DDEBF7"),  # blue
        "❌": PatternFill("solid", fgColor="FFC7CE"),  # red
        "-":  PatternFill("solid", fgColor="E2EFDA"),  # light green (already done)
    }

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data rows
    for r, row_data in enumerate(ROWS, start=2):
        priority = row_data[5]
        fill = fills.get(priority, None)
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = left if col in (1, 2, 7) else center
            if fill:
                cell.fill = fill

    # Column widths
    widths = {1: 32, 2: 28, 3: 14, 4: 8, 5: 18, 6: 10, 7: 50}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # ── Notes / Legend sheet ──
    ws2 = wb.create_sheet("Legend and Notes")
    legend = [
        "VSR Comparison Candidates — Legend",
        "",
        "Direct compare:",
        "  ★★★ = Strong direct comparison candidate (priority 🥇)",
        "  ★★  = Recommended if time permits",
        "  ★   = Possible direct comparison",
        "  △   = Marginal, depends on setup",
        "  ✗   = Not appropriate for direct comparison",
        "  ✗✗  = Definitely skip",
        "  DONE = Already evaluated",
        "  OURS = Our method",
        "",
        "Cite (related work value):",
        "  ★★★ = Must cite (close domain/method)",
        "  ★★  = Recommended cite",
        "  ★   = Optional cite",
        "  ✗   = Skip",
        "",
        "Priority colors:",
        "  🥇 (green)  = Highest priority for direct comparison",
        "  🥈 (yellow) = If time permits",
        "  🥉 (blue)   = Cite as related work, no direct comparison",
        "  ❌ (red)    = Skip entirely",
        "  - (light)   = Already done",
        "",
        "Recommended action plan:",
        "  1. Try DGAF-VSR (latest, github tszssong/DGAF-VSR) — direct comparison",
        "  2. Add RVRT — strong feed-forward baseline (3h)",
        "  3. Add MGLD-VSR or PatchVSR — diffusion VSR comparison",
        "  4. Cite all 🥉 papers in Related Work",
        "  5. Skip all ❌ papers entirely",
        "",
        "Setup compatibility:",
        "  BIx4 ✓        = Standard bicubic ×4 downsampling, your setup",
        "  Real-world    = Compressed/realistic degradation, different setup",
        "  Compressed    = H.264/H.265 compressed, different setup",
        "  Streaming     = Low-latency online VSR, different focus",
        "  Image SR      = Single image, not video",
        "  Different     = Different input modality or problem",
        "",
        "Filtering rule of thumb:",
        "  - If 'Setup compat.' = 'BIx4' → Direct comparison candidate",
        "  - If real-world/compressed → Cite only",
        "  - If different problem → Skip or skip + minor cite",
    ]
    for i, line in enumerate(legend, start=1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        if i in (3, 13, 19, 26, 33, 41):
            cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 80

    out = "comparison_candidates.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  Sheet 1: 'Comparison Candidates' ({len(ROWS)} methods)")
    print(f"  Sheet 2: 'Legend / Notes'")


if __name__ == '__main__':
    main()
