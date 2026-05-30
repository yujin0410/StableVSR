"""Middle-frame evaluation for Vimeo-90K-T.

Vimeo-90K-T provides GT only for the middle frame (`im4.png`), so the standard
per-frame loop in `eval.py` cannot be used as-is. This script:

* Reads the flattened recon layout `rec_path/<outer>_<inner>/im4.png`.
* Reads the 3-level GT layout `gt_path/<outer>/<inner>/im4.png`.
* Computes per-frame metrics (PSNR/SSIM/LPIPS/DISTS/MUSIQ/NIQE/CLIP-IQA).
* Optionally appends a single results row to a shared Excel workbook so that
  different models can be compared side-by-side.
"""

from torchmetrics.image.lpip import LearnedPerceptualImagePatchSimilarity as LPIPS
from torchmetrics.image import PeakSignalNoiseRatio as PSNR
from torchmetrics.image import StructuralSimilarityIndexMeasure as SSIM
import pyiqa
from DISTS_pytorch import DISTS
import os
import argparse
import warnings

import numpy as np
import torch
from PIL import Image
from torchvision.transforms import ToTensor
from tqdm import tqdm

warnings.filterwarnings("ignore")


def parse_args():
    parser = argparse.ArgumentParser(description="Middle-frame eval for Vimeo-90K-T.")
    parser.add_argument("--rec_path", type=str, required=True,
                        help="Folder with flattened recon: <outer>_<inner>/im{1..7}.png")
    parser.add_argument("--gt_path", type=str, required=True,
                        help="Vimeo target root: <outer>/<inner>/im4.png")
    parser.add_argument("--model_name", type=str, default="",
                        help="Model label used when appending to Excel.")
    parser.add_argument("--excel_path", type=str, default="",
                        help="If set, append a result row to this .xlsx file.")
    return parser.parse_args()


def build_metrics(device):
    return {
        "psnr":  PSNR(data_range=1).to(device),
        "ssim":  SSIM(data_range=1).to(device),
        "lpips": LPIPS(normalize=True).to(device),
        "dists": DISTS().to(device),
        "musiq": pyiqa.create_metric("musiq",   device="cuda", as_loss=False),
        "niqe":  pyiqa.create_metric("niqe",    device="cuda", as_loss=False),
        "clip":  pyiqa.create_metric("clipiqa", device="cuda", as_loss=False),
    }


def append_to_excel(path, model_name, dataset, means):
    """Append a single row of mean metrics to an .xlsx file (creating it if needed)."""
    from openpyxl import Workbook, load_workbook

    header = ["Model", "Dataset", "PSNR", "SSIM", "LPIPS",
              "DISTS", "MUSIQ", "NIQE", "CLIP-IQA", "N"]
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "vimeo_mid_frame"
        ws.append(header)

    ws.append([
        model_name, dataset,
        means["psnr"], means["ssim"], means["lpips"], means["dists"],
        means["musiq"], means["niqe"], means["clip"], means["n"],
    ])
    wb.save(path)


def main():
    args = parse_args()

    print("Run with arguments:")
    for k, v in vars(args).items():
        print(f"  {k}: {v}")

    device = torch.device("cuda")
    m = build_metrics(device)
    to_tensor = ToTensor()

    names = sorted(os.listdir(args.rec_path))
    pbar = tqdm(names, ncols=100)

    sums = {k: 0.0 for k in ("psnr", "ssim", "lpips", "dists", "musiq", "niqe", "clip")}
    n = 0
    missing = 0

    for name in pbar:
        if "_" not in name:
            continue
        outer, inner = name.split("_", 1)
        gt_p = os.path.join(args.gt_path, outer, inner, "im4.png")
        rec_p = os.path.join(args.rec_path, name, "im4.png")
        if not (os.path.exists(gt_p) and os.path.exists(rec_p)):
            missing += 1
            continue

        with torch.no_grad():
            gt = to_tensor(Image.open(gt_p)).unsqueeze(0).to(device)
            rec = to_tensor(Image.open(rec_p)).unsqueeze(0).to(device)

            sums["psnr"]  += m["psnr"](gt, rec).item()
            sums["ssim"]  += m["ssim"](gt, rec).item()
            sums["lpips"] += m["lpips"](gt, rec).item()
            sums["dists"] += m["dists"](gt, rec).item()
            sums["musiq"] += m["musiq"](rec).item()
            sums["niqe"]  += m["niqe"](rec).item()
            sums["clip"]  += m["clip"](rec).item()
        n += 1

    pbar.close()
    if n == 0:
        raise RuntimeError(f"No (rec, gt) pairs found under {args.rec_path}")

    means = {k: float(np.round(v / n, 4)) for k, v in sums.items()}
    means["n"] = n

    print(
        f"N={n} (missing={missing}) | "
        f"PSNR: {means['psnr']:.2f}, SSIM: {means['ssim']:.3f}, "
        f"LPIPS: {means['lpips']:.3f}, DISTS: {means['dists']:.3f}, "
        f"MUSIQ: {means['musiq']:.2f}, NIQE: {means['niqe']:.2f}, "
        f"CLIP-IQA: {means['clip']:.3f}"
    )

    if args.excel_path:
        append_to_excel(args.excel_path, args.model_name or "(unnamed)",
                        "Vimeo-90K-T", means)
        print(f"Appended row to {args.excel_path}")


if __name__ == "__main__":
    main()
