from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "VSR Comparison"

# Headers
headers = [
    "Model",
    "Year / Venue",
    "Foundation Model",
    "Trainable Params (M)",
    "Foundation Params (M)",
    "Total Inference Params (M)",
    "Training GPU Type",
    "Training GPU Count",
    "Training Duration",
    "Training Iterations / Batch",
    "Inference VRAM (paper)",
    "A6000x2 (48GBx2) Run?",
    "Code/Ckpt Released",
    "Notes",
]

# Data rows - leave blank where paper doesn't explicitly state
data = [
    [
        "BasicVSR++",
        "CVPR 2022",
        "(none, recurrent CNN)",
        "7.32",
        "—",
        "7.32",
        "",
        "",
        "",
        "300K iter (supp)",
        "",
        "Yes",
        "Yes",
        "Non-diffusion baseline",
    ],
    [
        "StableVSR (baseline)",
        "ECCV 2024",
        "Stable Diffusion 2.1",
        "~207",
        "~1,290",
        "~1,490",
        "",
        "",
        "",
        "",
        "~17 GB (training)",
        "Yes",
        "Yes",
        "Our direct baseline; ControlNet + TCM trained, UNet frozen",
    ],
    [
        "Ours",
        "2026 (in submission)",
        "Stable Diffusion 2.1",
        "213.92",
        "~1,290",
        "~1,510",
        "NVIDIA RTX A6000",
        "2",
        "(internal)",
        "30,000 iter",
        "(measured)",
        "Yes",
        "Internal",
        "ControlNet + dual-SFT (DT-CWT freq.) trained, UNet frozen",
    ],
    [
        "DGAF-VSR",
        "CVPR 2026 (arXiv 2025-11)",
        "Stable Diffusion 2.1",
        "~277 (ControlNet)",
        "~870",
        "~1,140",
        "",
        "",
        "",
        "",
        "",
        "Yes",
        "Yes",
        "Optical flow guided warping + dense feature guidance",
    ],
    [
        "FlashVSR",
        "CVPR 2026",
        "Wan2.1-1.3B",
        "~460 (LQ proj + TCDec + VAE)",
        "~1,420",
        "~1,880",
        "NVIDIA A100-80G",
        "32",
        "5 days (2+1+2 for 3 stages)",
        "Batch size 32",
        "",
        "Yes",
        "Yes",
        "Three-stage progressive distillation",
    ],
    [
        "DLoRAL",
        "NeurIPS 2025",
        "Stable Diffusion 2.1 + RAM-Swin",
        "~10.5 (LoRA rank 4)",
        "~1,800 (SD2.1 + RAM)",
        "~1,810",
        "",
        "",
        "",
        "",
        "",
        "Yes",
        "Yes",
        "Dual-LoRA (C-LoRA + D-LoRA), one-step",
    ],
    [
        "Vivid-VR",
        "ICLR 2026",
        "CogVideoX1.5-5B + CogVLM2",
        "~1,372 (controlnet + connectors)",
        "~12,000",
        "~13,400",
        "",
        "",
        "",
        "~10K (distill) + ~20K (long training)",
        "~43 GB (inference, 121 frames)",
        "Yes",
        "Yes",
        "Concept distillation from T2V foundation",
    ],
    [
        "STAR",
        "ICCV 2025",
        "CogVideoX-5B / I2VGen-XL",
        "~2,041 (one adapter file)",
        "~5,000 (CogVideoX-5B)",
        "~7,000",
        "",
        "",
        "",
        "200K text-video pairs (OpenVid-1M)",
        ">=24 GB (inference, ~39 GB for 4x 426x240 72f)",
        "Yes",
        "Yes",
        "Spatial-Temporal Augmentation with T2V model",
    ],
    [
        "DOVE",
        "NeurIPS 2025",
        "CogVideoX1.5-5B (fine-tuned)",
        "~5,570 (transformer fully tuned)",
        "(replaced)",
        "~7,500",
        "NVIDIA A100/A800-80G",
        "4",
        "(unstated)",
        "Total batch size 8, 2 stages",
        "",
        "Yes",
        "Yes",
        "One-step diffusion via full fine-tune of CogVideoX",
    ],
    [
        "SeedVR2",
        "ICLR 2026",
        "own DiT (no separate foundation)",
        "3,391 (full DiT)",
        "—",
        "3,391",
        "NVIDIA H100-80G",
        "72",
        "(unstated)",
        "~100 frames 720p per batch (seq+data parallel)",
        ">=80 GB (inference, 1x H100 for 720p 100f)",
        "OOM",
        "Yes",
        "One-step DiT via adversarial post-training",
    ],
    [
        "UltraVSR",
        "ACM MM 2025",
        "Stable Diffusion 2.1 + RAM-Swin",
        "~10.5 (LoRA rank 4)",
        "~1,800",
        "~1,810",
        "",
        "",
        "",
        "Input seq len 6, 512x512 crop",
        "",
        "OOM (Ours environment)",
        "Yes",
        "LoRA on VAE + UNet, RTS modules; one-step",
    ],
    [
        "Upscale-A-Video",
        "CVPR 2024",
        "Stable Diffusion x4 Upscaler",
        "~1,200 (all components)",
        "(integrated)",
        "~1,200",
        "",
        "",
        "",
        "WebVid10M (~335K) + YouHQ-Train (38,576 videos)",
        "A100 used for inference benchmark (512x512, 50 frames)",
        "OOM (Ours environment)",
        "Yes",
        "Temporal layers in U-Net+VAE + flow-guided propagation",
    ],
    [
        "MGLD-VSR",
        "ECCV 2024",
        "Stable Diffusion 2.1",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "A100 used for inference benchmark (128->512)",
        "Yes (likely)",
        "Yes",
        "Motion-guided latent diffusion",
    ],
]

ws.append(headers)
for row in data:
    ws.append(row)

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
ours_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
oom_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for row_idx in range(2, len(data) + 2):
    model_name = ws.cell(row=row_idx, column=1).value
    oom = ws.cell(row=row_idx, column=12).value
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
        if model_name == "Ours":
            cell.fill = ours_fill
        elif oom and "OOM" in str(oom):
            cell.fill = oom_fill

# Column widths
widths = {
    1: 22,  # Model
    2: 18,
    3: 28,
    4: 18,
    5: 18,
    6: 22,
    7: 22,
    8: 14,
    9: 22,
    10: 32,
    11: 30,
    12: 22,
    13: 18,
    14: 50,
}
for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 40

# Freeze header row
ws.freeze_panes = "A2"

out = "/home/user/StableVSR/VSR_comparison.xlsx"
wb.save(out)
print(f"Saved: {out}")
